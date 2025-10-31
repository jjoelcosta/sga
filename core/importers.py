# core/importers.py
from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime
from typing import IO, Tuple, List, Dict, Any

from django.db import transaction
from openpyxl import load_workbook

from .models import Empresa, Colaborador, Veiculo, Atribuicao

# Helpers simples / normalizadores
_DIGITS_RE = re.compile(r"\D+")


def _only_digits(s: str) -> str:
    return _DIGITS_RE.sub("", (s or ""))


def _norm_str(s: str) -> str:
    return (s or "").strip()


def _norm_cpf(s: str) -> str:
    cpf = _only_digits(s)
    return cpf if len(cpf) == 11 else ""


def _norm_cnpj(s: str) -> str:
    cnpj = _only_digits(s)
    return cnpj if len(cnpj) == 14 else ""


def _norm_placa(s: str) -> str:
    if not s:
        return ""
    placa = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    return placa if 7 <= len(placa) <= 8 else ""


def _read_xlsx_bytes(content: bytes):
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [(_norm_str(h) or f"col_{idx}") for idx, h in enumerate(row)]
            continue
        rows.append(
            {headers[j]: (row[j] if row and j < len(row) else None) for j in range(len(headers))}
        )
    return rows


def _read_csv_bytes(content: bytes):
    # tenta utf-8, se falhar tenta latin1
    try:
        text = content.decode("utf-8")
    except Exception:
        text = content.decode("latin1")
    reader = csv.DictReader(text.splitlines())
    return [{k.strip(): (v or "").strip() for k, v in row.items()} for row in reader]


def _open_rows(uploaded_file) -> List[Dict[str, Any]]:
    name = (uploaded_file.name or "").lower()
    content = uploaded_file.read()
    if name.endswith(".xlsx"):
        return _read_xlsx_bytes(content)
    if name.endswith(".csv"):
        return _read_csv_bytes(content)
    raise ValueError("Formato não suportado, use CSV ou XLSX.")


def _parse_date(s: str):
    s = _norm_str(s or "")
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


# Funções específicas de importação: retornam (criados_ou_atualizados, erros_list)
@transaction.atomic
def importar_empresas(uploaded_file) -> Tuple[int, List[str]]:
    rows = _open_rows(uploaded_file)
    ok = 0
    erros: List[str] = []
    for i, r in enumerate(rows, start=2):
        cnpj_raw = r.get("CNPJ") or r.get("cnpj") or r.get("Cnpj") or ""
        cnpj = _norm_cnpj(cnpj_raw)
        nome = _norm_str(r.get("Nome") or r.get("nome") or "")
        responsavel = _norm_str(r.get("Responsavel") or r.get("responsavel") or "")
        telefone = _norm_str(r.get("Telefone") or r.get("telefone") or "")

        if not cnpj or not nome:
            erros.append(f"Linha {i}: faltando CNPJ válido ou Nome")
            continue

        Empresa.objects.update_or_create(
            cnpj=cnpj,
            defaults={"nome": nome, "responsavel": responsavel, "telefone": telefone},
        )
        ok += 1
    return ok, erros


@transaction.atomic
def importar_colaboradores(uploaded_file) -> Tuple[int, List[str]]:
    rows = _open_rows(uploaded_file)
    ok = 0
    erros: List[str] = []
    for i, r in enumerate(rows, start=2):
        cpf_raw = r.get("CPF") or r.get("cpf") or ""
        cpf = _norm_cpf(cpf_raw)
        nome = _norm_str(r.get("Nome") or r.get("nome") or "")
        funcao = _norm_str(r.get("Funcao") or r.get("função") or r.get("funcao") or "")
        emp_cnpj_raw = r.get("Empresa_CNPJ") or r.get("empresa_cnpj") or r.get("EmpresaCNPJ") or ""
        emp_cnpj = _norm_cnpj(emp_cnpj_raw)

        if not cpf or not nome or not emp_cnpj:
            erros.append(f"Linha {i}: faltando CPF válido, Nome ou Empresa_CNPJ")
            continue

        try:
            empresa = Empresa.objects.get(cnpj=emp_cnpj)
        except Empresa.DoesNotExist:
            erros.append(f"Linha {i}: Empresa {emp_cnpj} não encontrada")
            continue

        Colaborador.objects.update_or_create(
            cpf=cpf,
            defaults={"nome": nome, "funcao": funcao, "empresa": empresa},
        )
        ok += 1
    return ok, erros


@transaction.atomic
def importar_veiculos(uploaded_file) -> Tuple[int, List[str]]:
    rows = _open_rows(uploaded_file)
    ok = 0
    erros: List[str] = []
    for i, r in enumerate(rows, start=2):
        placa_raw = r.get("Placa") or r.get("placa") or ""
        placa = _norm_placa(placa_raw)
        marca = _norm_str(r.get("Marca") or r.get("marca") or "")
        modelo = _norm_str(r.get("Modelo") or r.get("modelo") or "")
        cor = _norm_str(r.get("Cor") or r.get("cor") or "")
        cpf_raw = r.get("CPF") or r.get("cpf") or ""
        cpf = _norm_cpf(cpf_raw)

        if not placa:
            erros.append(f"Linha {i}: faltando Placa válida")
            continue

        colab = None
        if cpf:
            colab = Colaborador.objects.filter(cpf=cpf).first()
            if not colab:
                erros.append(f"Linha {i}: Colaborador CPF {cpf} não encontrado, veículo será criado sem vínculo")

        Veiculo.objects.update_or_create(
            placa=placa,
            defaults={"marca": marca, "modelo": modelo, "cor": cor, "colaborador": colab},
        )
        ok += 1
    return ok, erros


@transaction.atomic
def importar_atribuicoes(uploaded_file) -> Tuple[int, List[str]]:
    rows = _open_rows(uploaded_file)
    ok = 0
    erros: List[str] = []
    for i, r in enumerate(rows, start=2):
        cpf_raw = r.get("CPF") or r.get("cpf") or ""
        cpf = _norm_cpf(cpf_raw)
        placa_raw = r.get("Placa") or r.get("placa") or ""
        placa = _norm_placa(placa_raw)
        local = _norm_str(r.get("Local") or r.get("local") or "")
        di = _norm_str(r.get("DataInicio") or r.get("data_inicio") or "")
        df = _norm_str(r.get("DataFim") or r.get("data_fim") or "")
        status = _norm_str(r.get("Status") or r.get("status") or "PENDENTE").upper() or "PENDENTE"

        if not cpf or not local or not di:
            erros.append(f"Linha {i}: faltando CPF válido, Local ou DataInicio")
            continue

        colab = Colaborador.objects.filter(cpf=cpf).first()
        if not colab:
            erros.append(f"Linha {i}: Colaborador CPF {cpf} não encontrado")
            continue

        veic = None
        if placa:
            veic = Veiculo.objects.filter(placa=placa).first()
            if not veic:
                erros.append(f"Linha {i}: Veículo placa {placa} não encontrado, atribuição seguirá sem veículo")

        d_inicio = _parse_date(di)
        d_fim = _parse_date(df) if df else None
        if not d_inicio:
            erros.append(f"Linha {i}: DataInicio inválida")
            continue

        Atribuicao.objects.create(
            colaborador=colab,
            veiculo=veic,
            local=local,
            data_inicio=d_inicio,
            data_fim=d_fim,
            status=status if status in {"LIBERADO", "PENDENTE", "BLOQUEADO"} else "PENDENTE",
        )
        ok += 1
    return ok, erros


# Função de entrada unificada para a view
def importar_arquivo(file_obj, tipo: str) -> Dict[str, Any]:
    """
    Recebe o arquivo (UploadedFile) e o tipo: 'empresa', 'colaborador', 'veiculo', 'atribuicao'.
    Retorna dict: {"ok": bool, "criados_ou_atualizados": int, "erros": [...], "detalhes": {...}}
    """
    tipo = (tipo or "").lower()
    try:
        if tipo == "empresa":
            ok, erros = importar_empresas(file_obj)
        elif tipo == "colaborador":
            ok, erros = importar_colaboradores(file_obj)
        elif tipo == "veiculo":
            ok, erros = importar_veiculos(file_obj)
        elif tipo == "atribuicao":
            ok, erros = importar_atribuicoes(file_obj)
        else:
            return {"ok": False, "criados_ou_atualizados": 0, "erros": [f"Tipo desconhecido: {tipo}"], "detalhes": {}}

        return {
            "ok": len(erros) == 0,
            "criados_ou_atualizados": ok,
            "erros": erros,
            "detalhes": {"tipo": tipo},
        }
    except Exception as exc:
        # Em caso de exceção não esperada, capturamos e retornamos erro para a view salvar no LoteImportacao
        return {"ok": False, "criados_ou_atualizados": 0, "erros": [f"Erro interno: {exc}"], "detalhes": {}}